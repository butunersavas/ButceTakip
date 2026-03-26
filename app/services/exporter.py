import logging
import csv
import io
import calendar
from datetime import date
from typing import Literal

from fastapi import Response
from openpyxl import Workbook
from sqlalchemy import func
from sqlmodel import Session, select

from app.models import BudgetItem, Expense, ExpenseStatus, PlanEntry, Scenario
from app.services.analytics import compute_quarterly_summary
from app.schemas import PurchaseFormPreparedReportItem


CURRENCY_SYMBOL = "$"
logger = logging.getLogger(__name__)

EXPORT_COLUMN_DEFINITIONS: list[tuple[str, str]] = [
    ("type", "type"),
    ("budget_code", "budget_code"),
    ("budget_name", "budget_name"),
    ("scenario", "scenario"),
    ("year", "year"),
    ("month", "month"),
    ("amount", "amount"),
    ("date", "date"),
    ("quantity", "quantity"),
    ("unit_price", "unit_price"),
    ("vendor", "vendor"),
    ("description", "description"),
    ("department", "Departman"),
    ("status", "status"),
    ("out_of_budget", "out_of_budget"),
    ("capex_opex", "capex_opex"),
    ("asset_type", "asset_type"),
    ("record_owner", "kaydi_giren_kullanici"),
]
DEFAULT_EXPORT_COLUMN_KEYS = [key for key, _ in EXPORT_COLUMN_DEFINITIONS]
EXPORT_COLUMN_HEADER_MAP = dict(EXPORT_COLUMN_DEFINITIONS)


def _format_currency(value: float) -> str:
    amount = 0 if value is None else float(value)
    formatted = f"{amount:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return f"{formatted}{CURRENCY_SYMBOL}"


def _normalize_columns(columns: list[str] | None) -> list[str]:
    if not columns:
        return DEFAULT_EXPORT_COLUMN_KEYS
    valid = [column for column in columns if column in EXPORT_COLUMN_HEADER_MAP]
    return valid or DEFAULT_EXPORT_COLUMN_KEYS


def _resolve_date_bounds(
    year: int,
    month: int | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
) -> tuple[date, date]:
    if start_date and end_date:
        if start_date > end_date:
            raise ValueError("Başlangıç tarihi bitiş tarihinden büyük olamaz.")
        return start_date, end_date
    if start_date and not end_date:
        return start_date, date(year, 12, 31)
    if end_date and not start_date:
        return date(year, 1, 1), end_date
    if month is not None:
        return date(year, month, 1), date(year, month, calendar.monthrange(year, month)[1])
    return date(year, 1, 1), date(year, 12, 31)


def _resolve_month_bounds(
    year: int,
    month: int | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
) -> tuple[int, int] | None:
    if start_date or end_date or month is not None:
        start, end = _resolve_date_bounds(year, month, start_date, end_date)
        return start.month, end.month
    return None


def _get_expenses(
    session: Session,
    year: int,
    scenario_id: int | None = None,
    budget_item_id: int | None = None,
    month: int | None = None,
    department: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
) -> list[Expense]:
    date_start, date_end = _resolve_date_bounds(year, month, start_date, end_date)
    query = select(Expense).where(
        Expense.expense_date >= date_start,
        Expense.expense_date <= date_end,
    )
    if scenario_id is not None:
        query = query.where(Expense.scenario_id == scenario_id)
    if budget_item_id is not None:
        query = query.where(Expense.budget_item_id == budget_item_id)
    if department is not None:
        department_items = select(PlanEntry.budget_item_id).where(
            PlanEntry.year == year,
            PlanEntry.department == department,
        )
        if scenario_id is not None:
            department_items = department_items.where(PlanEntry.scenario_id == scenario_id)
        query = query.where(Expense.budget_item_id.in_(department_items))
    query = query.order_by(Expense.expense_date)
    return session.exec(query).all()


def _get_budget_item_map(session: Session) -> dict[int, BudgetItem]:
    return {item.id: item for item in session.exec(select(BudgetItem)).all()}


def _get_scenario_map(session: Session) -> dict[int, Scenario]:
    return {scenario.id: scenario for scenario in session.exec(select(Scenario)).all()}


def _append_plan_rows(
    sheet,
    plans: list[PlanEntry],
    budget_items: dict[int, BudgetItem],
    scenarios: dict[int, Scenario],
    columns: list[str],
):
    for plan in plans:
        budget_item = budget_items.get(plan.budget_item_id)
        scenario = scenarios.get(plan.scenario_id) if plan.scenario_id else None
        row_map = {
            "type": "plan",
            "budget_code": budget_item.code if budget_item else "",
            "budget_name": budget_item.name if budget_item else "",
            "scenario": scenario.name if scenario else "",
            "year": plan.year,
            "month": plan.month,
            "amount": _format_currency(plan.amount),
            "date": "",
            "quantity": "",
            "unit_price": "",
            "vendor": "",
            "description": "",
            "department": plan.department or "",
            "status": "",
            "out_of_budget": "false",
            "capex_opex": budget_item.map_category if budget_item and budget_item.map_category else "",
            "asset_type": budget_item.map_attribute if budget_item and budget_item.map_attribute else "",
            "record_owner": "",
        }
        sheet.append([row_map.get(column, "") for column in columns])


def _append_expense_rows(
    sheet,
    expenses: list[Expense],
    budget_items: dict[int, BudgetItem],
    scenarios: dict[int, Scenario],
    columns: list[str],
):
    for expense in expenses:
        budget_item = budget_items.get(expense.budget_item_id)
        scenario = scenarios.get(expense.scenario_id) if expense.scenario_id else None
        row_map = {
            "type": "expense",
            "budget_code": budget_item.code if budget_item else "",
            "budget_name": budget_item.name if budget_item else "",
            "scenario": scenario.name if scenario else "",
            "year": expense.expense_date.year,
            "month": expense.expense_date.month,
            "amount": _format_currency(expense.amount),
            "date": expense.expense_date.isoformat(),
            "quantity": expense.quantity,
            "unit_price": _format_currency(expense.unit_price),
            "vendor": expense.vendor or "",
            "description": expense.description or "",
            "department": "",
            "status": expense.status.value,
            "out_of_budget": "true" if expense.is_out_of_budget else "false",
            "capex_opex": budget_item.map_category if budget_item and budget_item.map_category else "",
            "asset_type": budget_item.map_attribute if budget_item and budget_item.map_attribute else "",
            "record_owner": expense.kaydi_giren_kullanici or "",
        }
        sheet.append([row_map.get(column, "") for column in columns])


def get_purchase_forms_prepared(
    session: Session,
    year: int,
    scenario_id: int | None = None,
    month: int | None = None,
    department: str | None = None,
    budget_item_id: int | None = None,
    capex_opex: str | None = None,
) -> list[PurchaseFormPreparedReportItem]:
    query = select(PlanEntry, BudgetItem).join(BudgetItem, PlanEntry.budget_item_id == BudgetItem.id).where(
        PlanEntry.year == year,
        PlanEntry.purchase_requested.is_(True),
    )

    if scenario_id is not None:
        query = query.where(PlanEntry.scenario_id == scenario_id)
    if month is not None:
        query = query.where(PlanEntry.month == month)
    if department is not None:
        query = query.where(PlanEntry.department == department)
    if budget_item_id is not None:
        query = query.where(PlanEntry.budget_item_id == budget_item_id)
    if capex_opex:
        query = query.where(func.lower(func.coalesce(BudgetItem.map_category, "")) == capex_opex.lower())

    rows = session.exec(query.order_by(PlanEntry.month, BudgetItem.code)).all()
    return [
        PurchaseFormPreparedReportItem(
            budget_item_id=budget_item.id,
            budget_code=budget_item.code,
            budget_name=budget_item.name or budget_item.code,
            year=plan.year,
            month=plan.month,
            scenario_id=plan.scenario_id,
            department=plan.department or None,
            amount=plan.amount,
            capex_opex=budget_item.map_category,
            purchase_requested_at=plan.purchase_requested_at,
        )
        for plan, budget_item in rows
    ]


def export_csv(
    session: Session,
    year: int,
    scenario_id: int | None = None,
    budget_item_id: int | None = None,
) -> Response:
    output = io.StringIO()
    writer = csv.writer(output)
    budget_items = _get_budget_item_map(session)
    writer.writerow(
        [
            "type",
            "budget_item_id",
            "map_category",
            "map_attribute",
            "scenario_id",
            "year",
            "month",
            "amount",
            "currency",
            "status",
            "is_out_of_budget",
            "expense_date",
            "vendor",
            "description",
            "client_hostname",
            "kaydi_giren_kullanici",
        ]
    )

    plan_query = select(PlanEntry).where(PlanEntry.year == year)
    if scenario_id is not None:
        plan_query = plan_query.where(PlanEntry.scenario_id == scenario_id)
    if budget_item_id is not None:
        plan_query = plan_query.where(PlanEntry.budget_item_id == budget_item_id)
    for plan in session.exec(plan_query).all():
        budget_item = budget_items.get(plan.budget_item_id)
        map_attribute = budget_item.map_attribute if budget_item else ""
        map_category = budget_item.map_category if budget_item else ""
        writer.writerow(
            [
                "plan",
                plan.budget_item_id,
                map_category or "",
                map_attribute or "",
                plan.scenario_id,
                plan.year,
                plan.month,
                _format_currency(plan.amount),
                "USD",
                "",
                "",
                "",
                "",
                "",
            ]
        )

    expenses = _get_expenses(session, year, scenario_id, budget_item_id)
    for expense in expenses:
        budget_item = budget_items.get(expense.budget_item_id)
        map_attribute = budget_item.map_attribute if budget_item else ""
        map_category = budget_item.map_category if budget_item else ""
        writer.writerow(
            [
                "expense",
                expense.budget_item_id,
                map_category or "",
                map_attribute or "",
                expense.scenario_id,
                expense.expense_date.year,
                expense.expense_date.month,
                _format_currency(expense.amount),
                "USD",
                expense.status.value,
                str(expense.is_out_of_budget).lower(),
                expense.expense_date.isoformat(),
                expense.vendor or "",
                expense.description or "",
                expense.client_hostname or "",
                expense.kaydi_giren_kullanici or "",
            ]
        )
    response = Response(content=output.getvalue(), media_type="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=budget_export.csv"
    return response


def export_xlsx(
    session: Session,
    year: int,
    scenario_id: int | None = None,
    budget_item_id: int | None = None,
    month: int | None = None,
    department: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    columns: list[str] | None = None,
) -> Response:
    wb = Workbook()
    plan_sheet = wb.active
    plan_sheet.title = "BudgetData"
    selected_columns = _normalize_columns(columns)
    plan_sheet.append([EXPORT_COLUMN_HEADER_MAP[column] for column in selected_columns])
    budget_items = _get_budget_item_map(session)
    scenarios = _get_scenario_map(session)
    plan_query = select(PlanEntry).where(PlanEntry.year == year)
    if scenario_id is not None:
        plan_query = plan_query.where(PlanEntry.scenario_id == scenario_id)
    if budget_item_id is not None:
        plan_query = plan_query.where(PlanEntry.budget_item_id == budget_item_id)
    if department is not None:
        plan_query = plan_query.where(PlanEntry.department == department)
    month_bounds = _resolve_month_bounds(year, month, start_date, end_date)
    if month_bounds:
        plan_query = plan_query.where(PlanEntry.month >= month_bounds[0], PlanEntry.month <= month_bounds[1])
    plans = session.exec(plan_query).all()
    _append_plan_rows(plan_sheet, plans, budget_items, scenarios, selected_columns)

    expenses = _get_expenses(
        session, year, scenario_id, budget_item_id, month, department, start_date, end_date
    )
    _append_expense_rows(plan_sheet, expenses, budget_items, scenarios, selected_columns)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("export_debug export_type=%s year=%s scenario=%s budget_item=%s row_count=%s", "budget_xlsx", year, scenario_id, budget_item_id, len(plans) + len(expenses))
    response = Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = "attachment; filename=budget_export.xlsx"
    return response


def export_filtered_expenses_xlsx(
    session: Session,
    year: int,
    scenario_id: int | None = None,
    budget_item_id: int | None = None,
    month: int | None = None,
    department: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    columns: list[str] | None = None,
    *,
    filter_type: Literal["out_of_budget", "cancelled"],
) -> Response:
    expenses = _get_expenses(
        session, year, scenario_id, budget_item_id, month, department, start_date, end_date
    )
    budget_items = _get_budget_item_map(session)
    scenarios = _get_scenario_map(session)
    if filter_type == "out_of_budget":
        filtered = [expense for expense in expenses if expense.is_out_of_budget]
        sheet_title = "OutOfBudgetExpenses"
        filename = f"out_of_budget_expenses_{year}.xlsx"
    else:
        filtered = [expense for expense in expenses if expense.status == ExpenseStatus.CANCELLED]
        sheet_title = "CancelledExpenses"
        filename = f"cancelled_expenses_{year}.xlsx"

    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_title
    selected_columns = _normalize_columns(columns)
    sheet.append([EXPORT_COLUMN_HEADER_MAP[column] for column in selected_columns])
    _append_expense_rows(sheet, filtered, budget_items, scenarios, selected_columns)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("export_debug export_type=%s year=%s scenario=%s budget_item=%s filter_type=%s row_count=%s", "filtered_expenses_xlsx", year, scenario_id, budget_item_id, filter_type, len(filtered))
    response = Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def export_purchase_forms_prepared_xlsx(
    session: Session,
    year: int,
    scenario_id: int | None = None,
    month: int | None = None,
    department: str | None = None,
    budget_item_id: int | None = None,
    capex_opex: str | None = None,
) -> Response:
    items = get_purchase_forms_prepared(session, year, scenario_id, month, department, budget_item_id, capex_opex)
    wb = Workbook()
    sheet = wb.active
    sheet.title = "PreparedPurchaseForms"
    sheet.append([
        "Budget Code",
        "Budget Name",
        "Year",
        "Month",
        "Scenario ID",
        "Department",
        "Capex/Opex",
        "Amount",
        "Purchase Requested At",
    ])

    for item in items:
        sheet.append(
            [
                item.budget_code,
                item.budget_name,
                item.year,
                item.month,
                item.scenario_id or "",
                item.department or "",
                item.capex_opex or "",
                _format_currency(item.amount),
                item.purchase_requested_at.isoformat() if item.purchase_requested_at else "",
            ]
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("export_debug export_type=%s year=%s month=%s scenario=%s department=%s budget_item=%s capex_opex=%s row_count=%s", "purchase_forms_prepared_xlsx", year, month, scenario_id, department, budget_item_id, capex_opex, len(items))
    response = Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = (
        f"attachment; filename=purchase_forms_prepared_{year}.xlsx"
    )
    return response


def export_quarterly_csv(
    session: Session,
    year: int,
    scenario_id: int | None = None,
    budget_item_id: int | None = None,
) -> Response:
    summary = compute_quarterly_summary(session, year, scenario_id, budget_item_id)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "quarter",
            "planned",
            "actual",
            "saving",
            "out_of_budget",
            "cancelled",
        ]
    )
    for entry in summary:
        writer.writerow(
            [
                f"Q{entry.quarter}",
                _format_currency(entry.planned),
                _format_currency(entry.actual),
                _format_currency(entry.saving),
                _format_currency(entry.out_of_budget),
                _format_currency(entry.cancelled),
            ]
        )
    response = Response(content=output.getvalue(), media_type="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=quarterly_summary.csv"
    return response


def export_quarterly_xlsx(
    session: Session,
    year: int,
    scenario_id: int | None = None,
    budget_item_id: int | None = None,
    month: int | None = None,
    department: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
) -> Response:
    summary = compute_quarterly_summary(session, year, scenario_id, budget_item_id, department)
    if month is not None:
        summary = [
            entry
            for entry in summary
            if ((entry.quarter - 1) * 3 + 1) <= month <= (entry.quarter * 3)
        ]
    if start_date and end_date:
        allowed_quarters = {
            ((m - 1) // 3) + 1
            for m in range(start_date.month, end_date.month + 1)
        }
        summary = [entry for entry in summary if entry.quarter in allowed_quarters]
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Quarterly Summary"
    sheet.append([
        "Quarter",
        "Planned (USD)",
        "Actual (USD)",
        "Saving (USD)",
        "Out of Budget (USD)",
        "Cancelled (USD)",
    ])
    for entry in summary:
        sheet.append(
            [
                f"Q{entry.quarter}",
                _format_currency(entry.planned),
                _format_currency(entry.actual),
                _format_currency(entry.saving),
                _format_currency(entry.out_of_budget),
                _format_currency(entry.cancelled),
            ]
        )
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("export_debug export_type=%s year=%s scenario=%s budget_item=%s row_count=%s", "quarterly_xlsx", year, scenario_id, budget_item_id, len(summary))
    response = Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = "attachment; filename=quarterly_summary.xlsx"
    return response


def get_export_preview_summary(
    session: Session,
    year: int,
    scenario_id: int | None = None,
    budget_item_id: int | None = None,
    month: int | None = None,
    department: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
) -> dict[str, float | int]:
    plan_query = select(PlanEntry).where(PlanEntry.year == year)
    if scenario_id is not None:
        plan_query = plan_query.where(PlanEntry.scenario_id == scenario_id)
    if budget_item_id is not None:
        plan_query = plan_query.where(PlanEntry.budget_item_id == budget_item_id)
    if department is not None:
        plan_query = plan_query.where(PlanEntry.department == department)
    month_bounds = _resolve_month_bounds(year, month, start_date, end_date)
    if month_bounds:
        plan_query = plan_query.where(PlanEntry.month >= month_bounds[0], PlanEntry.month <= month_bounds[1])
    plans = session.exec(plan_query).all()

    expenses = _get_expenses(
        session, year, scenario_id, budget_item_id, month, department, start_date, end_date
    )
    planned = float(sum(plan.amount for plan in plans))
    actual = float(sum(expense.amount for expense in expenses if expense.status == ExpenseStatus.RECORDED and not expense.is_out_of_budget))
    out_of_budget = float(sum(expense.amount for expense in expenses if expense.status == ExpenseStatus.RECORDED and expense.is_out_of_budget))
    cancelled = float(sum(expense.amount for expense in expenses if expense.status == ExpenseStatus.CANCELLED))
    return {
        "total_plan": planned,
        "total_actual": actual,
        "total_out_of_budget": out_of_budget,
        "total_cancelled": cancelled,
        "record_count": len(plans) + len(expenses),
        "quarterly_data_count": len(plans) + len(expenses),
        "out_of_budget_count": len([expense for expense in expenses if expense.is_out_of_budget]),
        "cancelled_count": len([expense for expense in expenses if expense.status == ExpenseStatus.CANCELLED]),
    }
