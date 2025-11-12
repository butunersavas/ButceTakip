import csv
import io
import json
import re
import unicodedata
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

from fastapi import HTTPException, UploadFile, status
from sqlmodel import Session, select

from app.models import BudgetItem, Expense, PlanEntry, Scenario
from app.schemas import ImportSummary

MONTH_ALIASES = {
    "jan": 1,
    "feb": 2,
    "mart": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
    "oca": 1,
    "şub": 2,
    "sub": 2,
    "nis": 4,
    "mayıs": 5,
    "mayis": 5,
    "haz": 6,
    "tem": 7,
    "ağu": 8,
    "agu": 8,
    "eyl": 9,
    "eki": 10,
    "kas": 11,
    "ara": 12,
}

EXPENSE_COLUMN_KEYWORDS = {"actual", "expense", "spend", "spent", "harcama", "gerçekleşen"}


def _ensure_budget_item(
    session: Session,
    code: str,
    name: str | None = None,
    map_attribute: str | None = None,
    capex_opex: str | None = None,
    asset_type: str | None = None,
) -> BudgetItem:
    item = session.exec(select(BudgetItem).where(BudgetItem.code == code)).first()
    if item:
        updated = False
        if name and item.name != name:
            item.name = name
            updated = True
        if map_attribute and item.map_attribute != map_attribute:
            item.map_attribute = map_attribute
            updated = True
        if capex_opex and item.capex_opex != capex_opex:
            item.capex_opex = capex_opex
            updated = True
        if asset_type and item.asset_type != asset_type:
            item.asset_type = asset_type
            updated = True
        if updated:
            session.add(item)
            session.commit()
            session.refresh(item)
        return item
    if not name:
        raise HTTPException(status_code=400, detail=f"Missing name for budget item {code}")
    item = BudgetItem(
        code=code,
        name=name,
        map_attribute=map_attribute,
        capex_opex=capex_opex,
        asset_type=asset_type,
    )
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


def _extract_map_attribute(data: dict[str, Any]) -> str | None:
    for key in ("map_attribute", "map nitelik", "map-nitelik", "mapnitelik", "map_nitelik"):
        value = data.get(key)
        if value is None:
            continue
        if isinstance(value, str):
            stripped = value.strip()
        else:
            stripped = str(value).strip()
        if stripped:
            return stripped
    return None


def _normalize_optional_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
    else:
        text = str(value).strip()
    return text or None


def _extract_text_with_aliases(data: dict[str, Any], *aliases: str) -> str | None:
    if not aliases:
        return None
    normalized_map: dict[str, str] = {}
    for key in data.keys():
        normalized = re.sub(r"[^a-z0-9]", "", key.lower())
        normalized_map[normalized] = key
    for alias in aliases:
        normalized_alias = re.sub(r"[^a-z0-9]", "", alias.lower())
        source_key = normalized_map.get(normalized_alias)
        if source_key is None:
            continue
        candidate = _normalize_optional_str(data.get(source_key))
        if candidate:
            return candidate
    return None


def _normalize_capex_opex(value: Any) -> str | None:
    text = _normalize_optional_str(value)
    if not text:
        return None
    normalized = text.upper()
    if "CAPEX" in normalized:
        return "CAPEX"
    if "OPEX" in normalized:
        return "OPEX"
    return text


def _extract_capex_opex(data: dict[str, Any]) -> str | None:
    value = _extract_text_with_aliases(
        data,
        "capex_opex",
        "capex opex",
        "capex-opex",
        "capex/opex",
        "capexopex",
    )
    return _normalize_capex_opex(value)


def _normalize_asset_type(value: Any) -> str | None:
    text = _normalize_optional_str(value)
    if not text:
        return None
    return text


def _extract_asset_type(data: dict[str, Any]) -> str | None:
    return _normalize_asset_type(
        _extract_text_with_aliases(data, "asset_type", "asset type", "varlik tipi", "varlik_tipi", "varliktipi")
    )


def _normalize_code(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = ascii_text.strip().upper()
    ascii_text = re.sub(r"[^A-Z0-9]+", "_", ascii_text)
    ascii_text = ascii_text.strip("_")
    if not ascii_text:
        ascii_text = "ITEM"
    return ascii_text[:64]


def _parse_month_header(header: str) -> tuple[int, int] | None:
    normalized = header.strip().lower()
    if not normalized or "total" in normalized or "genel toplam" in normalized:
        return None
    # remove words that can precede month info such as plan/actual
    normalized = normalized.replace("plan", " ")
    normalized = normalized.replace("budget", " ")
    normalized = normalized.replace("gerçekleşen", " ")
    normalized = normalized.replace("actual", " ")
    normalized = normalized.replace("expense", " ")
    normalized = normalized.replace("harcama", " ")
    normalized = re.sub(r"\s+", " ", normalized)

    month_match = None
    for month_key, month_value in MONTH_ALIASES.items():
        if month_key in normalized:
            month_match = (month_key, month_value)
            break
    if not month_match:
        return None
    year_match = re.search(r"(20\d{2}|19\d{2}|\d{2})", normalized)
    if not year_match:
        return None
    raw_year = int(year_match.group(1))
    if raw_year < 100:
        raw_year += 2000 if raw_year < 70 else 1900
    return (month_match[1], raw_year)


def _find_header_index(headers: list[str], *candidates: str) -> int | None:
    lowered = [header.lower() for header in headers]
    for candidate in candidates:
        candidate_lower = candidate.lower()
        if candidate_lower in lowered:
            return lowered.index(candidate_lower)
    for index, header in enumerate(lowered):
        if any(candidate_lower in header for candidate_lower in map(str.lower, candidates)):
            return index
    return None


def _coerce_str(value: Any, field: str) -> str:
    if value is None:
        raise ValueError(f"Missing value for {field}")
    if isinstance(value, str):
        text = value.strip()
    else:
        text = str(value).strip()
    if not text:
        raise ValueError(f"Missing value for {field}")
    return text


def _coerce_int(value: Any, field: str) -> int:
    if value is None:
        raise ValueError(f"Missing value for {field}")
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        raise ValueError(f"Missing value for {field}")
    return int(text)


def _coerce_float(value: Any, field: str) -> float:
    if value is None:
        raise ValueError(f"Missing value for {field}")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        raise ValueError(f"Missing value for {field}")
    return float(text)


def _coerce_date(value: Any, field: str) -> date:
    if value is None:
        raise ValueError(f"Missing value for {field}")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        raise ValueError(f"Missing value for {field}")
    return date.fromisoformat(text)


def _get_value(data: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in data and data[key] not in (None, ""):
            return data[key]
    return None


def import_json(file: UploadFile, session: Session) -> ImportSummary:
    try:
        payload = json.loads(file.file.read().decode("utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid JSON file") from exc

    summary = ImportSummary()
    if isinstance(payload, dict) and "plans" in payload:
        summary.imported_plans += _import_plan_list(payload["plans"], session)
    elif isinstance(payload, dict) and "year" in payload and "items" in payload:
        summary.imported_plans += _import_year_month_structure(payload, session)
    elif isinstance(payload, list):
        summary.imported_plans += _import_plan_list(payload, session)
    else:
        raise HTTPException(status_code=400, detail="Unsupported JSON schema")
    summary.message = "JSON import completed"
    return summary


def _import_plan_list(data: list[dict], session: Session) -> int:
    imported = 0
    for entry in data:
        try:
            map_attribute = _extract_map_attribute(entry)
            capex_opex = _extract_capex_opex(entry)
            asset_type = _extract_asset_type(entry)
            item = _ensure_budget_item(
                session,
                entry["budget_code"],
                entry.get("budget_name"),
                map_attribute,
                capex_opex,
                asset_type,
            )
            scenario = _get_or_create_scenario(
                session, entry.get("scenario"), int(entry.get("year"))
            )
            plan = PlanEntry(
                year=int(entry["year"]),
                month=int(entry["month"]),
                amount=float(entry["amount"]),
                scenario_id=scenario.id,
                budget_item_id=item.id,
            )
            session.add(plan)
            session.commit()
            imported += 1
        except Exception:
            session.rollback()
    return imported


def _import_year_month_structure(data: dict, session: Session) -> int:
    imported = 0
    year = int(data["year"])
    scenario = _get_or_create_scenario(session, data.get("scenario"), year)
    for item_code, months in data.get("items", {}).items():
        map_attribute = _extract_map_attribute(months)
        capex_opex = _extract_capex_opex(months)
        asset_type = _extract_asset_type(months)
        item = _ensure_budget_item(
            session,
            item_code,
            months.get("name"),
            map_attribute,
            capex_opex,
            asset_type,
        )
        for month_str, amount in months.get("plan", {}).items():
            plan = PlanEntry(
                year=year,
                month=int(month_str),
                amount=float(amount),
                scenario_id=scenario.id,
                budget_item_id=item.id,
            )
            session.add(plan)
            session.commit()
            imported += 1
    return imported


def _get_or_create_scenario(session: Session, scenario_name: str | None, year: int) -> Scenario:
    if scenario_name:
        scenario = session.exec(
            select(Scenario).where(Scenario.name == scenario_name).where(Scenario.year == year)
        ).first()
        if scenario:
            return scenario
        scenario = Scenario(name=scenario_name, year=year)
        session.add(scenario)
        session.commit()
        session.refresh(scenario)
        return scenario
    scenario = session.exec(select(Scenario).where(Scenario.year == year)).first()
    if scenario:
        return scenario
    scenario = Scenario(name=f"Default {year}", year=year)
    session.add(scenario)
    session.commit()
    session.refresh(scenario)
    return scenario


def _import_pivot_style_rows(
    data_rows: list[tuple[Any, ...]],
    headers_raw: list[str],
    session: Session,
    summary: ImportSummary,
) -> bool:
    normalized_headers = [header.lower() for header in headers_raw]
    if not any("row labels" in header or "row label" in header for header in normalized_headers):
        return False

    month_columns: list[tuple[int, int, int]] = []
    for index, header in enumerate(headers_raw):
        parsed = _parse_month_header(header)
        if not parsed:
            continue
        header_lower = header.lower()
        if any(keyword in header_lower for keyword in EXPENSE_COLUMN_KEYWORDS):
            continue
        month_columns.append((index, parsed[0], parsed[1]))

    if not month_columns:
        return False

    name_index = _find_header_index(headers_raw, "row labels", "row label", "kalem")
    if name_index is None:
        return False

    code_index = _find_header_index(headers_raw, "budget_code", "budget code", "kod", "code")
    map_index = _find_header_index(headers_raw, "map attribute", "map_attribute", "map nitelik", "map")
    scenario_index = _find_header_index(headers_raw, "scenario", "senaryo", "bench")
    type_index = _find_header_index(headers_raw, "type", "tip", "type export")
    capex_index = _find_header_index(
        headers_raw,
        "capex_opex",
        "capex opex",
        "capex-opex",
        "capex/opex",
        "capexopex",
    )
    asset_index = _find_header_index(
        headers_raw,
        "asset_type",
        "asset type",
        "varlık tipi",
        "varlik tipi",
        "varlik_tipi",
    )

    for row in data_rows:
        if not row or name_index >= len(row):
            continue
        raw_name = row[name_index]
        if raw_name is None:
            continue
        if isinstance(raw_name, str):
            name = raw_name.strip()
        else:
            name = str(raw_name).strip()
        if not name:
            continue
        if "total" in name.lower() or "genel toplam" in name.lower():
            continue

        map_attribute_value: str | None = None
        if map_index is not None and map_index < len(row):
            map_attribute_value = _normalize_optional_str(row[map_index])

        capex_opex_value: str | None = None
        if capex_index is not None and capex_index < len(row):
            capex_opex_value = _normalize_capex_opex(row[capex_index])

        asset_type_value: str | None = None
        if asset_index is not None and asset_index < len(row):
            asset_type_value = _normalize_asset_type(row[asset_index])

        code_value: str | None = None
        if code_index is not None and code_index < len(row):
            code_cell = row[code_index]
            if code_cell not in (None, ""):
                code_value = str(code_cell).strip()
        if not code_value:
            code_value = map_attribute_value or name
        budget_code = _normalize_code(code_value)

        scenario_value: str | None = None
        if scenario_index is not None and scenario_index < len(row):
            scenario_cell = row[scenario_index]
            if scenario_cell not in (None, ""):
                scenario_value = str(scenario_cell).strip()
        entry_type_value = "plan"
        if type_index is not None and type_index < len(row):
            type_cell = row[type_index]
            if type_cell not in (None, ""):
                entry_type_value = str(type_cell).strip().lower()
        if entry_type_value not in {"plan", "planlama", "budget"}:
            # Currently only plan rows are supported
            continue

        item = _ensure_budget_item(
            session,
            budget_code,
            name,
            map_attribute_value,
            capex_opex_value,
            asset_type_value,
        )

        for index, month, year in month_columns:
            if index >= len(row):
                continue
            amount = row[index]
            if amount in (None, "", 0):
                continue
            try:
                amount_value = float(amount)
            except (TypeError, ValueError):
                continue
            scenario = _get_or_create_scenario(session, scenario_value, year)
            plan = PlanEntry(
                year=year,
                month=month,
                amount=amount_value,
                scenario_id=scenario.id,
                budget_item_id=item.id,
            )
            session.add(plan)
            session.commit()
            summary.imported_plans += 1

    return True


def import_csv(file: UploadFile, session: Session) -> ImportSummary:
    try:
        content = file.file.read().decode("utf-8")
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail="CSV must be UTF-8 encoded") from exc
    reader = csv.DictReader(io.StringIO(content))
    summary = ImportSummary()
    for row in reader:
        try:
            entry_type = row.get("type", "plan").lower()
            map_attribute = _extract_map_attribute(row)
            capex_opex = _extract_capex_opex(row)
            asset_type = _extract_asset_type(row)
            if entry_type == "plan":
                item = _ensure_budget_item(
                    session,
                    row["budget_code"],
                    row.get("budget_name"),
                    map_attribute,
                    capex_opex,
                    asset_type,
                )
                scenario = _get_or_create_scenario(session, row.get("scenario"), int(row["year"]))
                plan = PlanEntry(
                    year=int(row["year"]),
                    month=int(row["month"]),
                    amount=float(row["amount"]),
                    scenario_id=scenario.id,
                    budget_item_id=item.id,
                )
                session.add(plan)
                session.commit()
                summary.imported_plans += 1
            elif entry_type == "expense":
                item = _ensure_budget_item(
                    session,
                    row["budget_code"],
                    row.get("budget_name"),
                    map_attribute,
                    capex_opex,
                    asset_type,
                )
                scenario = _get_or_create_scenario(session, row.get("scenario"), int(row["year"]))
                expense = Expense(
                    budget_item_id=item.id,
                    scenario_id=scenario.id,
                    expense_date=date.fromisoformat(row["date"]),
                    amount=float(row["amount"]),
                    quantity=float(row.get("quantity", 1) or 1),
                    unit_price=float(row.get("unit_price", 0) or 0),
                    vendor=row.get("vendor"),
                    description=row.get("description"),
                    is_out_of_budget=row.get("out_of_budget", "false").lower() == "true",
                )
                session.add(expense)
                session.commit()
                summary.imported_expenses += 1
            else:
                summary.skipped_rows += 1
        except Exception:
            session.rollback()
            summary.skipped_rows += 1
    summary.message = "CSV import completed"
    return summary


def import_xlsx(file: UploadFile, session: Session) -> ImportSummary:
    try:
        file.file.seek(0)
        workbook = load_workbook(file.file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail="Invalid XLSX file") from exc

    sheet = workbook.active
    if sheet is None:
        raise HTTPException(status_code=400, detail="XLSX file does not contain any sheets")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="XLSX file is empty")

    headers_raw: list[str] = []
    headers: list[str] = []
    for cell in rows[0]:
        if cell is None:
            headers_raw.append("")
            headers.append("")
        elif isinstance(cell, str):
            stripped = cell.strip()
            headers_raw.append(stripped)
            headers.append(stripped.lower())
        else:
            stripped = str(cell).strip()
            headers_raw.append(stripped)
            headers.append(stripped.lower())

    summary = ImportSummary()
    if _import_pivot_style_rows(rows[1:], headers_raw, session, summary):
        summary.message = "XLSX import completed"
        return summary
    for values in rows[1:]:
        row: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[index] if index < len(values) else None
        if not row:
            continue
        entry_type = str(row.get("type") or "plan").lower()
        map_attribute = _extract_map_attribute(row)
        capex_opex = _extract_capex_opex(row)
        asset_type = _extract_asset_type(row)
        try:
            budget_code = _coerce_str(
                _get_value(row, "budget_code", "budget code", "kod"),
                "budget_code",
            )
            budget_name = _get_value(row, "budget_name", "budget name", "ad")
            scenario_name = _get_value(row, "scenario", "senaryo")
            year_value = _coerce_int(_get_value(row, "year", "yıl"), "year")

            if entry_type == "plan":
                month_value = _coerce_int(_get_value(row, "month", "ay"), "month")
                amount_value = _coerce_float(_get_value(row, "amount", "tutar"), "amount")
                item = _ensure_budget_item(
                    session,
                    budget_code,
                    budget_name,
                    map_attribute,
                    capex_opex,
                    asset_type,
                )
                scenario = _get_or_create_scenario(session, scenario_name, year_value)
                plan = PlanEntry(
                    year=year_value,
                    month=month_value,
                    amount=amount_value,
                    scenario_id=scenario.id,
                    budget_item_id=item.id,
                )
                session.add(plan)
                session.commit()
                summary.imported_plans += 1
            elif entry_type == "expense":
                item = _ensure_budget_item(
                    session,
                    budget_code,
                    budget_name,
                    map_attribute,
                    capex_opex,
                    asset_type,
                )
                scenario = _get_or_create_scenario(session, scenario_name, year_value)
                amount_value = _coerce_float(_get_value(row, "amount", "tutar"), "amount")
                quantity_value = _get_value(row, "quantity", "adet")
                unit_price_value = _get_value(row, "unit_price", "birim fiyat")
                date_value = _get_value(row, "date", "tarih")
                expense = Expense(
                    budget_item_id=item.id,
                    scenario_id=scenario.id,
                    expense_date=_coerce_date(date_value, "date"),
                    amount=amount_value,
                    quantity=float(quantity_value) if quantity_value not in (None, "") else 1.0,
                    unit_price=float(unit_price_value) if unit_price_value not in (None, "") else 0.0,
                    vendor=_get_value(row, "vendor", "satıcı"),
                    description=_get_value(row, "description", "açıklama"),
                    is_out_of_budget=str(_get_value(row, "out_of_budget", "bütçe_dışı") or "false").lower()
                    == "true",
                )
                session.add(expense)
                session.commit()
                summary.imported_expenses += 1
            else:
                summary.skipped_rows += 1
        except Exception:
            session.rollback()
            summary.skipped_rows += 1

    summary.message = "XLSX import completed"
    return summary
